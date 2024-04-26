import tkinter as tk
from tkinter import *
from tkinter import messagebox, filedialog
from tkinter.messagebox import askyesno

from tkinter import ttk
from tkcalendar import Calendar

from logica_guardar_datos import evaluar_guardar_archivos, evaluar_guardar_medidores_factor
from graficos_consumos import grafico_perfiles_instrumentacion
import datetime
from datetime import datetime
import os

import openpyxl
# Base de datos
from db import SessionLocal, SessionLocal2
from modelos import DatosMedidorConsumo, Medidores, DatosMedidorInstrumentacion, Permitidos
from sqlalchemy import func

# obtener el numero serial del dispositivo donde se pretende ejecutar el aplicativo
import subprocess
output = subprocess.check_output('wmic bios get serialnumber').decode("utf-8")
# se devuelve un encabezado en la parte superior y tras un salto de línea se muestra el número serial
# se separa en un array para tomarlo como segundo elemento
lines = output.strip().split('\n')
serial_number = lines[1]

# reportes
# primer grafico y data
from reporte_consumos import obtener_consumo_por_medidor_y_rango

# función para obtener información específica de los medidores, esta información serpa visible en la pestaña principal "Datos histróricos"
def obtener_medidores_info():
    # se accede a la sesión de la base de datos
    db = SessionLocal()
    try:
        # Obtener los IDs únicos de los medidores que tienen registros en DatosMedidorConsumo
        medidores_con_consumo = db.query(DatosMedidorConsumo.meter_id).distinct().all()

        # lista que cargará la data de los medidores según la base de dato
        medidores_info = []

        # Iterar sobre los IDs de los medidores con consumo
        for medidor_id, in medidores_con_consumo:
            # Verificar si el medidor también existe en la tabla Medidores
            medidor_info = db.query(Medidores).filter_by(id=medidor_id).first()

            if medidor_info:
                # Obtener el primer y último registro de consumo para cada medidor
                primer_registro_consumo = db.query(func.min(DatosMedidorConsumo.date)).filter_by(meter_id=medidor_id).scalar()
                ultimo_registro_consumo = db.query(func.max(DatosMedidorConsumo.date)).filter_by(meter_id=medidor_id).scalar()
                # Calcular el número de huecos para cada medidor
                huecos = db.query(DatosMedidorConsumo).filter(DatosMedidorConsumo.meter_id == medidor_id, DatosMedidorConsumo.kwh_del.is_(None)).count()
                # Agregar los datos del medidor y su consumo a la lista
                medidores_info.append((medidor_id, medidor_info.sed ,primer_registro_consumo, ultimo_registro_consumo, huecos, medidor_info.marca, medidor_info.factor))

        return medidores_info
    
    finally:
        db.close()

# función para eliminar los registros del medidor, interactuando con la interfaz principal del proyecto
def eliminar_registros_medidor(medidor):
    # se accede a la sesión de la base de datos
    db = SessionLocal()
    try:
        # consulta para eliminar el medidor o los medidores seleccionados
        db.query(DatosMedidorConsumo).filter_by(meter_id=medidor).delete()
        db.query(DatosMedidorInstrumentacion).filter_by(meter_id=medidor).delete()
        db.commit()
        messagebox.showinfo("Eliminación exitosa", f"Los datos del medidor {medidor} han sido eliminados.")
    except Exception as e:
        # se deshace la eliminación
        db.rollback()
        messagebox.showerror("Error", f"No se pudo eliminar los datos del medidor {medidor}: {str(e)}")
    finally:
        db.close()

''' LAS DOS FUNCIONES SIGUIENTES SERÁN UTILIZADAS PARA IMPORTAR DATOS '''
# esta funcion sirve para los perfiles de carga
def importar_lecturas():
    # declara variable vacía

    # tipos de archivos permitidos
    tipos_archivos = [("Archivos CSV y PRN", "*.csv;*.prn")]
    # ventana de selección
    archivos_seleccionados = filedialog.askopenfilenames(title="Seleccione archivos CSV o PRN - Perfiles de Carga", filetypes=tipos_archivos)

    # Obtener los nombres de los archivos seleccionados
    nombres_archivos = [os.path.basename(archivo) for archivo in archivos_seleccionados]
    mensaje_confirmacion = "¿Desea importar los siguientes archivos?\n\n"
    mensaje_confirmacion+="\n".join(nombres_archivos)

    '''AGREGAR AQUÍ MENSAJE DE CONFIRMACIÓN (NOMBRES/CANTIDAD DE ARCHIVOS A IMPORTAR)'''
    if archivos_seleccionados:
        aviso_confirmacion = askyesno(title='Confirmación', message=mensaje_confirmacion)

        if aviso_confirmacion==True:
            # se ejecuta la función 'evaluar_guardar_archivos' dentro de una variable, la función creada en el archivo 'logica_guardar_datos.py' devuelve un una lista de nombres
            # esta lista de nombres estará guardada en la variable 'archivos_no_validos'
            archivos_no_validos = evaluar_guardar_archivos(archivos_seleccionados)
            # si existen datos en esta variable, se mostrará el nombre de estos archivos en un mensaje
            if archivos_no_validos:
                mensaje = "Los siguientes archivos no son válidos:\n" + "\n".join(archivos_no_validos)
                messagebox.showwarning("Archivos no válidos", mensaje)
            else:
            # si la lista esta vacía por no existir archivos inválidos se mostrará un mensaje de confirmación
                messagebox.showinfo("Éxito", "Todos los archivos fueron importados correctamente.")
                # se ejecuta la siguiente función que refresca la información mostrada en la interfaz
                actualizar_lista_medidores()
        else:
            messagebox.showinfo("Aviso", "Operación Detenida")
    else:
        messagebox.showinfo("Mensaje", "No se seleccionaron archivos para importar.")

# esta función sirve para importar información sobre los medidores y el factor
'''modificar en base a lo requerido'''
def importar_medidores():
    # tipos de archivos permitidos
    tipos_archivos = [("Excel", "*xlsx")]
    # ventana de selección
    archivos_seleccionados = filedialog.askopenfilenames(title="Seleccione el archivo correcto", filetypes=tipos_archivos)

    # Obtener el nombre del archivo seleccionado.
    nombre_archivo = [os.path.basename(archivo) for archivo in archivos_seleccionados]
    mensaje_confirmacion = "¿Desea importar el siguiente archivo?\n\n"
    mensaje_confirmacion+="\n".join(nombre_archivo)

    if len(archivos_seleccionados)==1:
        aviso_confirmacion = askyesno(title='Confirmación', message=mensaje_confirmacion)

        if aviso_confirmacion==True:
            # se ejecuta la función 'medidores_factor' dentro de una variable, la función creada en el archivo 'logica_guardar_datos.py' devuelve un una lista de nombres
            # esta lista de nombres estará guardada en la variable 'archivos_no_validos'
            archivo_no_valido = evaluar_guardar_medidores_factor(archivos_seleccionados)

            if archivo_no_valido!=0:
                mensaje = "Archivo inválido."
                messagebox.showwarning("Error al seleccionar el archivo", mensaje)
            else:
            # si la lista esta vacía por no existir archivos inválidos se mostrará un mensaje de confirmación
                messagebox.showinfo("Éxito", "Guardado correcto.")
                # se ejecuta la siguiente función que refresca la información mostrada en la interfaz
                actualizar_lista_medidores()
    elif len(archivos_seleccionados)<1:
        messagebox.showinfo("Mensaje", "Operación cancelada")
    else:
        messagebox.showerror("Mensaje", "Debe seleccionar el archivo excel específico")

'''Comienza la sección para generar reportes'''
# primera funcion para generar reportes
def generar_reportes_one():
    if not treeview_medidores.selection():
        messagebox.showwarning("Medidor no seleccionado", "Por favor, seleccione un medidor antes de generar el reporte.")
        return
    else:
        abrir_ventana_periodo()
    #print(lista_medidores)

# se abre la ventana para seleccionar periodo. (para que esta ventana aparezca, primero se debe de haber seleccionado uno o varios medidores)
def abrir_ventana_periodo():
    global ventana_periodo
    global seleccion_fecha_inicio
    global seleccion_fecha_fin
    ventana_periodo = tk.Toplevel()
    ventana_periodo.title("Periodo")

    ventana_periodo.resizable(0,0)
    
    # label y selección de fecha de inicio
    label_fecha_inicio = ttk.Label(ventana_periodo, text="Fecha de inicio:")
    label_fecha_inicio.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    seleccion_fecha_inicio = Calendar(ventana_periodo)
    seleccion_fecha_inicio.grid(row=0, column=1, padx=5, pady=5)

    # label y selección de fecha de fin
    label_fecha_fin = ttk.Label(ventana_periodo, text="Fecha de fin:")
    label_fecha_fin.grid(row=1, column=0, padx=5, pady=5, sticky="w")
    seleccion_fecha_fin = Calendar(ventana_periodo)
    seleccion_fecha_fin.grid(row=1, column=1, padx=5, pady=5)

    # Botón de aceptar generar el reporte
    boton_aceptar = ttk.Button(ventana_periodo, text="Aceptar", command=generar_reportes_periodo)
    boton_aceptar.grid(row=2, column=0, columnspan=2, pady=10)

    ventana_periodo.mainloop()

# función que toma las fechas para generar el reporte
def generar_reportes_periodo():
    global fecha_inicio
    global fecha_fin
    # fechas
    fecha_inicio_str = seleccion_fecha_inicio.get_date()
    fecha_fin_str = seleccion_fecha_fin.get_date()

    # Convertir las cadenas a objetos de fecha
    fecha_inicio = datetime.strptime(fecha_inicio_str, "%m/%d/%y").date()
    fecha_fin = datetime.strptime(fecha_fin_str, "%m/%d/%y").date()
    # se cierra la nueva ventana tras haber seleccionado las fechas de manera exitosa
    ventana_periodo.destroy()
    # se ejecuta la última función, se pasan las fechas como argmentos

    generar_reportes(fecha_inicio, fecha_fin)

    ''' Luego de haber generado los reportes en la interfaz, esta sección mostrará en un mensaje de alerta el id de los medidores que no tengan registros en todas las fechas
    seleccionadas; sin embargo de la misma manera muestra el reporte ya que al menos cuenta con data en los días iniciales'''
    id_medidores_fechas_excedentes = ""
    # con esta condición se evalúa si existe una cantidad de datos en el arreglo que contiene a los medidores cuyas de selección se exceden a las de registro
    if(len(datos_medidores_fechas)>=1):
        for dato_fecha in datos_medidores_fechas:
            id_medidores_fechas_excedentes+= f" - {dato_fecha} "

        messagebox.showinfo("Medidores sin registros en el rango de fechas seleccionadas", id_medidores_fechas_excedentes)

# se crea una variable global para luego poder usar la información que devuelva la función 'generar_reportes'
datos_obtenidos_globales = []

# se crea esta variable global, TENDRÁ LOS MEDIDORES CUYAS FECHAS DE SOLICITUD A LAS FECHAS DE REGISTRO
datos_medidores_fechas = ""

# función generar reportes, usando las fechas como arugmentos para otra función de se encuentra en 'reporte_cosumos'
def generar_reportes(fecha_inicio, fecha_fin):

    global datos_obtenidos_globales
    global datos_medidores_fechas

    # obtiene los identificadores de los medidores seleccionados
    medidores_seleccionados = [treeview_medidores.item(item, "text") for item in treeview_medidores.selection()]
    #print("DATA A MODO DE REPORTE")
    datos_obtenidos = []
    
    # por cada medidor seleccionado se obtiene su id de la lista para poder hacer uso de la función 'obtener_consumo_por_medidor_y_rango'
    # función que se encuentra en el archivo 'reporte_consumos.py'
    for medidor_id in medidores_seleccionados:
        # llamada a la función en reporte_consumos para obtener los datos de consumo en el rango de fechas especificado
        datos_consumo = obtener_consumo_por_medidor_y_rango(fecha_inicio, fecha_fin, medidor_id)
        #print(datos_consumo)
        # se agrega en la lista todas listas devuelvas al ejecutar la función en el archivo de reporte
        datos_obtenidos.append(datos_consumo)

    #print(datos_obtenidos)

    # iterando la lista de datos
    try:
        for datos in datos_obtenidos:
            # se inserta cada conjunto de datos como una nueva fila en el treeview_reportes
            medidor = datos[0]
            dia = datos[1]
            mes = datos[2]
            cant_v = datos[3]
            desde = datos[4]
            hasta = datos[5]
            acum = datos[6]
            ener_mes = datos[7]
            hora = datos[8]
            tipo = datos[9]
            treeview_reportes.insert("", "end", text=medidor, values=(dia, mes, cant_v, desde, hasta, acum, ener_mes, hora, tipo))

        datos_medidores_fechas = (datos_obtenidos[-1][-1])
            # se guarda en la variable global la lista de datos obtenidos
        datos_obtenidos_globales = datos_obtenidos
    except:
        messagebox.showwarning("Aviso", "No se pudo completar la operación, inténtelo nuevamente y revise los rangos de fecha seleccionados")

    '''el usuario prefiere que todos los gráficos sean guardados en una misma carpeta, uno detrás del otro
    para ello ya se cuenta con la tabla de medidores donde se encuentra el factor:
    lo primero sería calcular el promedio de las corrientes, se obtiene una gráfica de ese promedio
    tercero sería usar cada dato de ese promedio multiplicado por el factor, luego de ello se obtiene la segunda gráfica
    '''

# función para verificar si debe generarse el excel
def verificar_y_generar_excel():
    # revisa si existe contenido en el treeview_reportes
    if treeview_reportes.get_children():
        generar_excel()
    else:
        messagebox.showwarning("Sin datos", "No hay datos para generar el archivo Excel.")

# función para crear el excel con la data obtenida desde lo mostrado en el treeview de tkinter
def generar_excel():
    try:
        # ruta de la carpeta de descargas de Windows
        carpeta_descargas = os.path.join(os.path.expanduser('~'), 'Downloads')

        # ruta de la carpeta
        carpeta_aplicativo_reportes = os.path.join(carpeta_descargas, 'APLICATIVO_ENEL_SUMINISTROS_REPORTES')

        if not os.path.exists(carpeta_aplicativo_reportes):
           os.makedirs(carpeta_aplicativo_reportes)

        # ruta del archivo "DATOS_REPORTES.xlsx"
        archivo_excel = os.path.join(carpeta_descargas, carpeta_aplicativo_reportes, 'DATOS_REPORTES.xlsx')

        if os.path.exists(archivo_excel):
            # Si el archivo ya existe, abrirlo y cargar el libro de trabajo
            workbook = openpyxl.load_workbook(archivo_excel)
            # Seleccionar la hoja activa del libro
            sheet = workbook.active
        else:
            # Si el archivo no existe, crear uno nuevo y agregar una hoja llamada "Reporte"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reporte"

            # Agregar los encabezados a la primera fila
            headers = ["Medidor", "Dias", "Mes", "Cant. Vacíos", "Desde", "Hasta", "Acumulado", "Energía - Mes", "Hora max.Demanda", "Tipo de Consumo"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

        # Recorrer cada fila del treeview
        for row_id in treeview_reportes.get_children():
            # Obtener los valores de cada columna en la fila actual
            valores_fila = [treeview_reportes.item(row_id, "text")] + [treeview_reportes.set(row_id, col) for col in treeview_reportes["columns"]]
            # Agregar los valores de cada fila al archivo excel en la siguiente fila disponible
            sheet.append(valores_fila)

        # Guardar el archivo excel
        workbook.save(archivo_excel)
        messagebox.showinfo("Éxito", f"Datos guardados en {archivo_excel}")
    except Exception as e:
        messagebox.showerror("Error", f"Se produjo un error al guardar los datos: {str(e)}")
    else:
        # se obtiene el id de los medidores seleccionados para generar sus gráficos respectivos
        ids_medidores = [treeview_reportes.item(row_id, "text") for row_id in treeview_reportes.get_children()]
        # ejecutar función para generar gráfico
        grafico_perfiles_instrumentacion(ids_medidores, fecha_inicio, fecha_fin)


# método para eliminar datos de los registros
# se puede eliminar uno o más medidores
def eliminar_datos():
    # guarda en una variable los elementos seleccionados
    indices_seleccionados = treeview_medidores.selection()
    # se valida que existan elemenetos seleccionados
    if not indices_seleccionados:
        messagebox.showwarning("Selección vacía", "Por favor, seleccione al menos un medidor para eliminar.")
        return
    # mensaje de confirmación para el usuario
    respuesta = messagebox.askquestion("Confirmar Eliminación", "¿Está seguro de que desea eliminar el registro seleccionado?")
    if respuesta == "yes":
        # se recorre la lista de elementos seleccionados
        for indice in indices_seleccionados:
            # se selecciona el valor del elemento seleccionado
            medidor = treeview_medidores.item(indice, "text")
            # se ejecuta la función con todos los valores extraídos de la lista
            eliminar_registros_medidor(medidor)
        # se ejecuta la función de actualizar para refrescar el frame
        actualizar_lista_medidores()

def refresh_treeview():
    # Elimina todas las filas del treeview excepto los encabezados
    for row_id in treeview_reportes.get_children():
        treeview_reportes.delete(row_id)

# cerrar el aplicativo
def salir():
    root.destroy()

# modifica el treeview y refresca el elemento
def actualizar_lista_medidores():
    # elimina todos los hijos del treeview
    treeview_medidores.delete(*treeview_medidores.get_children())
    # ejecuta la función para obtener datos
    medidores_info = obtener_medidores_info()
    # recorre la información obtenida
    for medidor_info in medidores_info:
        medidor = medidor_info[0]
        medidor_sed = medidor_info[1]
        primer_registro = medidor_info[2]
        ultimo_registro = medidor_info[3]
        huecos = medidor_info[4]
        marca = medidor_info[5]
        factor = medidor_info[6]
        # inserta la información en el treeview
        treeview_medidores.insert("", "end", text=medidor, values=(medidor_sed, primer_registro, ultimo_registro, huecos, marca, factor))

'''ventana del aplicativo'''
root = tk.Tk()
root.title("Aplicativo para la campaña de totalizadores")
root.geometry("1450x800")
root.resizable(False, False)
root.config(bg="white")
icono = tk.PhotoImage(file="./images/fav-icon-ti.png")
root.iconphoto(True, icono)

'''Inicia Menu'''
# menu lateral
menu_frame = tk.Frame(root, bg="#2f3640", width=100)
menu_frame.pack(side="left", fill="y", expand=False)

# imagen del logo de teching
logo = tk.PhotoImage(file="./images/logo_teching.png")

# se carga la iamgen del logo en un label
menu_titulo = tk.Label(menu_frame, image=logo ,compound=tk.LEFT, font=("Arial", 14), bg="#2f3640", fg="white")
menu_titulo.pack(pady=35)

# se cargan los íconos para los botones
icono1 = tk.PhotoImage(file="images/img1.png")
imagen1 = icono1

icono2 = tk.PhotoImage(file="images/img2.png")
imagen2 = icono2

icono3 = tk.PhotoImage(file="images/img3.png")
imagen3 = icono3

icono4 = tk.PhotoImage(file="images/img4.png")
imagen4 = icono4

icono5 = tk.PhotoImage(file="images/img5.png")
imagen5 = icono5

def boton2():
    messagebox.showinfo("En efecto, se ha cambiado el comando")

# se crean los botones y se coloca su respectivo texto e íconos
boton_importar = tk.Button(menu_frame, text="   Importar Perfiles  ", image=imagen1, compound=tk.LEFT, command=importar_lecturas, width=155, cursor="hand2")
boton_importar.pack(pady=25, padx=20)
boton_medidores = tk.Button(menu_frame, text="   Importar Medidores  ", image=imagen5, compound=tk.LEFT, command=importar_medidores, width=155, cursor="hand2")
boton_medidores.pack(pady=25, padx=20)
boton_reportes = tk.Button(menu_frame, text="   Gen. Reportes   ", image=imagen2, compound=tk.LEFT, command=generar_reportes_one, width=155, cursor="hand2")
boton_reportes.pack(pady=25, padx=20)
boton_eliminar = tk.Button(menu_frame, text="   Eliminar   ", image=imagen3, compound=tk.LEFT, command=eliminar_datos, width=155, cursor="hand2")
boton_eliminar.pack(pady=25, padx=20)
boton_salir = tk.Button(menu_frame, text="   Salir   ", image=imagen4, compound=tk.LEFT ,command=salir, width=155, cursor="hand2")
boton_salir.pack(pady=25, padx=20)
'''Finaliza Menu'''

# Crear el Notebook
'''TK NOTEBOOK - PANEL DE PESTAÑAS EN DONDE ESTARÁN ALOJADAS LAS PESTAÑAS: ("DATOS HISTÓRICOS", "REPORTES")'''
notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)

# frame para los datos históricos
datos_historicos_frame = tk.Frame(notebook, bg="#2f3640", padx=20, pady=20)
notebook.add(datos_historicos_frame, text="Datos Históricos")

# frame para los reportes
reportes_frame = tk.Frame(notebook, bg="#2f3640", padx=20, pady=20)
notebook.add(reportes_frame, text="Reportes")

''' /*/*/*/*/* INICIA CONTENIDO DE DATOS HISTORICOS '''
# se asigna un titulo al frame de datos historicos
datos_historicos_title = tk.Label(datos_historicos_frame, text="Datos Históricos", bg="#2f3640", fg="white", font=("Arial", 12))
datos_historicos_title.pack(side="top", fill="x")
datos_historicos_title.configure(anchor="w") 

#se crea el treeview que mostrará información sobre los medidores
treeview_medidores = ttk.Treeview(datos_historicos_frame, columns=("SED", "Primer Registro", "Último Registro", "Huecos/Vacios", "Marca", "Factor"))

# Configurar el ancho de las columnas para que se ajusten automáticamente al contenido
treeview_medidores.column("#0", width=100)  # Ancho de la columna "Medidor"
treeview_medidores.column("#1", width=100, stretch=True)  # Ancho de la columna "SED"
treeview_medidores.column("#2", width=150, stretch=True)  # Ancho de la columna "Primer Registro"
treeview_medidores.column("#3", width=150, stretch=True)  # Ancho de la columna "Último Registro"
treeview_medidores.column("#4", width=100, stretch=True)  # Ancho de la columna "Huecos/Vacios"
treeview_medidores.column("#5", width=100, stretch=True)  # Ancho de la columna "Marca"
treeview_medidores.column("#6", width=100, stretch=True)  # Ancho de la columna "Factor"

# Asignar los encabezados de las columnas
treeview_medidores.heading("#0", text="Medidor")
treeview_medidores.heading("#1", text="SED")
treeview_medidores.heading("#2", text="Primer Registro")
treeview_medidores.heading("#3", text="Último Registro")
treeview_medidores.heading("#4", text="Huecos/Vacios")
treeview_medidores.heading("#5", text="Marca")
treeview_medidores.heading("#6", text="Factor")

# Hacer que el Treeview se ajuste al tamaño de la ventana y se muestren todas las columnas visibles
treeview_medidores.pack(expand=True, fill="both")
''' /*/*/*/*/* TERMINA CONTENIDO DE DATOS HISTORICOS '''

''' /*/*/*/*/* INICIA CONTENIDO PARA REPORTES ESTADÍSTICOS '''
# se asigna un titulo al frame de reportes estadísticos
reportes_estadisticos_title = tk.Label(reportes_frame, bg="#2f3640", fg="white", font=("Arial", 12))
reportes_estadisticos_title.pack(side="top", fill="x")
# reportes_estadisticos_title.configure(anchor="w", text="Reportes - Perfiles de Carga")

# agregando botones a la pestaña de reportes
tk.Label(reportes_estadisticos_title, bg="#2f3640", fg="white", font=("Arial", 12), text="Reportes - Perfiles de Carga").grid(row=0, column=0, sticky=W, padx=(10, 800))


# este botón servirá para refrescar la ventana de reportes
boton_refrescar_reportes = ttk.Button(reportes_estadisticos_title, text="Limpiar", cursor="hand2", command=refresh_treeview)
boton_refrescar_reportes.grid(row=0, column=1, padx=5)

# este botón servirá para generar los reportes en excel haciendo uso de los datos obtenidos
boton_generar_documento = ttk.Button(reportes_estadisticos_title, text="G. Doc", cursor="hand2", command=verificar_y_generar_excel)
boton_generar_documento.grid(row=0, column=2, padx=5)

# función que servirá para ordenar la información del treeview en la pestaña de reportes
def sort_column(treeview, col, reverse):
    # obtiene los datos y convertirlos a números si es posible, para luego poder ordenarlos de mayor a menor haciendo clic en los encabezados
    data = []
    for child in treeview.get_children(''):
        val = treeview.set(child, col)
        try:
            val = float(val)
        except ValueError:
            pass  # si no se puede convertir a número, lo deja como está
        data.append((val, child))
    
    # ordena los datos
    data.sort(reverse=reverse)

    # mueve los ítems en el Treeview a sus nuevas posiciones según el ordenamiento
    for index, (val, child) in enumerate(data):
        treeview.move(child, '', index)

    # configura el encabezado de la columna para invertir el orden al hacer clic
    treeview.heading(col, command=lambda: sort_column(treeview, col, not reverse))

# inserción de información
treeview_reportes = ttk.Treeview(reportes_frame, columns=("Medidor", "Días", "Mes", "Cant. Vacíos", "Desde", "Hasta", "Acumulado", "Energía - Mes", "Hora max.Demanda", "Tipo de Consumo"))
treeview_reportes.column("#0", width=120, stretch=NO)
treeview_reportes.column("#1", width=60, stretch=NO)
treeview_reportes.column("#2", width=120, stretch=NO)
treeview_reportes.column("#3", width=120, stretch=NO)
treeview_reportes.column("#4", width=120, stretch=NO)
treeview_reportes.column("#5", width=120, stretch=NO)
treeview_reportes.column("#6", width=120, stretch=NO)
treeview_reportes.column("#7", width=120, stretch=NO)
treeview_reportes.column("#8", width=120, stretch=NO)
treeview_reportes.column("#9", width=180, stretch=NO)

treeview_reportes.heading("#0", text="Medidor")
treeview_reportes.heading("#1", text="Dias")
treeview_reportes.heading("#2", text="Mes")
treeview_reportes.heading("#3", text="Cant. Vacíos", command=lambda: sort_column(treeview_reportes, "col3", False))
treeview_reportes.heading("#4", text="Desde")
treeview_reportes.heading("#5", text="Hasta")
treeview_reportes.heading("#6", text="Acumulado", command=lambda: sort_column(treeview_reportes, "col6", False))
treeview_reportes.heading("#7", text="Energía - Mes", command=lambda: sort_column(treeview_reportes, "col7", False))
treeview_reportes.heading("#8", text="Hora max.Demanda", command=lambda: sort_column(treeview_reportes, "col8", False))
treeview_reportes.heading("#9", text="Tipo de Consumo")

treeview_reportes.pack(expand=True, fill="both")

''' /*/*/*/*/* TERMINA CONTENIDO PARA REPORTES ESTADÍSTICOS '''

# Esta función sirve para refrescar la lista de medidores mostrados en el frame de datos históricos
actualizar_lista_medidores()

notebook.pack(expand=True, fill="both")

try:
    # se valida que el aplicativo esté siendo ejecutado donde el codigo serial coincida con el de la pc donde se pretenda usar
    db2 = SessionLocal2()

    # consulta para obtener todos los datos guardados
    usuarios_permitidos = db2.query(Permitidos).all()

    serial_valido = False

    # comparar el numero de serie de la pc actual con alguno de los registrados
    for dato in usuarios_permitidos:
        nombre_usuario = dato.usuario
        serie_usuario = dato.serie

        if(serial_number==serie_usuario):
            serial_valido=True
        break

    # se cierra la conexión después de la validación
    db2.close()

    # se mantiene en ejecución el aplicativo siempre y cuando se hayan realizado las validaciones
    if(serial_valido):
        #print("Se puede ejecutar el aplicativo")
        root.mainloop()
    else:
        #si no se encuentra el número de serie del dispositivo en donde se esta ejecutando este aplicativo, dentro de la base de datos
        #no se podrá abrir y aparecerá el siguiente mensaje
        messagebox.showerror("Acceso Denegado", "Comuníquese con el área de TI para gestionar su solicitud de acceso.")

except KeyboardInterrupt:
    print("Aplicación interrumpida por el usuario.")
finally:
    root.destroy()